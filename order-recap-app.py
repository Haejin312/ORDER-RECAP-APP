import os, io, json, base64, datetime, re, time, traceback
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import anthropic
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

TEMPLATE_PATH = 'Carhartt order recap 양식.xlsx'

# ══════════════════════════════════════════════════
# 실제 양식 기반 정확한 Row/Column 상수
# ══════════════════════════════════════════════════

# 헤더 (모두 col 3 = C열에 값 입력)
ROW_SEWING       = 4
ROW_PRINTING     = 5
ROW_WASHING      = 6
ROW_PRESENTATION = 7

# Revision History: AC열(29)=Date, AE열(31)=Notes, row 9부터 시작
ROW_REVISION_START = 9
COL_REV_DATE       = 29   # AC
COL_REV_NOTES      = 31   # AE

# Fabric Combination: row 11~14
# A(1)=Combo, C(3)=Body Code, D(4)=Body Desc, G(7)=Trim Code, H(8)=Trim Desc
FABRIC_ROW_START  = 11
FABRIC_COL_COMBO  = 1   # A
FABRIC_COL_BCODE  = 3   # C
FABRIC_COL_BDESC  = 4   # D
FABRIC_COL_TCODE  = 7   # G
FABRIC_COL_TDESC  = 8   # H

# 스케치 이미지: row 3, 각 Style# 열
SKETCH_ROW  = 3
SKETCH_COLS = {0: 16, 1: 19, 2: 22, 3: 25}   # P, S, V, Y

# PO 정보 행 (각 슬롯의 홀수 열에 입력)
ROW_PO         = 17
ROW_ORDER_UNIT = 18
ROW_EX_FACTORY = 19
ROW_DELIVERY   = 20
ROW_STYLE      = 21
ROW_COLOR_CODE = 22
ROW_COLOR_NAME = 23
ROW_FABRIC_LBL = 24
ROW_SHIP_TO    = 25
ROW_SHIP_MODE  = 26

# 사이즈별 PCS 행 (홀수 열 = PCS, 짝수 열 = FOB)
SIZE_ROWS = {
    'XS':30,'S':31,'M':32,'L':33,'XL':34,'2XL':35,
    '3XL':36,'4XL':37,'5XL':38,'6XL':39,
    'LTLL':40,'XLTLL':41,'2XLTLL':42,'3XLTLL':43,
    '4XLTLL':44,'5XLTLL':45
}
ROW_TOTAL_PCS = 46   # =SUM(Cx30:Cx45) 공식 이미 존재

# PO 슬롯: 16개, 홀수 열 (3,5,7,...,33)
# 각 슬롯에서 PCS=홀수열, FOB=홀수열+1
PO_SLOT_COLS = [3 + i * 2 for i in range(16)]


# ══════════════════════════════════════════════════
# 헬퍼 함수
# ══════════════════════════════════════════════════

def parse_date(s):
    if not s:
        return None
    for fmt in ['%m/%d/%Y','%Y-%m-%d','%d/%m/%Y','%m/%d/%y']:
        try:
            return datetime.datetime.strptime(str(s).strip(), fmt).date()
        except Exception:
            pass
    return None


def normalize_fiber(desc):
    if not desc:
        return 'UNKNOWN'
    d = desc.upper()
    if '60' in d and ('COTT' in d or 'POLY' in d):
        return '60/40 COTTON/POLY'
    if '90' in d and ('COTT' in d or 'POLY' in d):
        return '90/10 COTTON/POLY'
    if '100' in d and 'POLY' in d and 'COTT' not in d:
        return '100% POLYESTER'
    if '100' in d or ('COTT' in d and 'POLY' not in d):
        return '100% COTTON'
    return d.strip()


def w(ws, row, col, value):
    """병합 셀 포함 안전 쓰기"""
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    ws.cell(row=mr.min_row, column=mr.min_col).value = value
                    return
        else:
            cell.value = value
    except Exception:
        pass


def get_client(req):
    key = req.headers.get('X-API-Key', '')
    return anthropic.Anthropic(api_key=key) if key else anthropic.Anthropic()


# ══════════════════════════════════════════════════
# Routes
# ══════════════════════════════════════════════════

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status':'ok','template':os.path.exists(TEMPLATE_PATH)})


@app.route('/parse-po', methods=['POST'])
def parse_po():
    client  = get_client(request)
    files   = request.files.getlist('files')
    results = []

    for f in files:
        raw = f.read()
        b64 = base64.b64encode(raw).decode()
        try:
            resp = client.messages.create(
                model='claude-sonnet-4-5',
                max_tokens=2500,
                messages=[{
                    'role':'user',
                    'content':[
                        {
                            'type':'document',
                            'source':{'type':'base64','media_type':'application/pdf','data':b64}
                        },
                        {
                            'type':'text',
                            'text':(
                                'Extract from this Carhartt PO and return ONLY valid JSON:\n'
                                '{\n'
                                '  "po_number": "",\n'
                                '  "style": "",\n'
                                '  "color_code": "",\n'
                                '  "order_unit": 0,\n'
                                '  "ex_factory_date": "",\n'
                                '  "delivery_date": "",\n'
                                '  "ship_to": "",\n'
                                '  "ship_mode": "",\n'
                                '  "fabric_desc": "",\n'
                                '  "sizes": {"XS":0,"S":0,"M":0,"L":0,"XL":0,"2XL":0,'
                                '"3XL":0,"4XL":0,"5XL":0,"6XL":0,'
                                '"LTLL":0,"XLTLL":0,"2XLTLL":0,"3XLTLL":0,"4XLTLL":0,"5XLTLL":0},\n'
                                '  "fob_prices": {"XS":0,"S":0,"M":0,"L":0,"XL":0,"2XL":0,'
                                '"3XL":0,"4XL":0,"5XL":0,"6XL":0,'
                                '"LTLL":0,"XLTLL":0,"2XLTLL":0,"3XLTLL":0,"4XLTLL":0,"5XLTLL":0},\n'
                                '  "total_qty": 0\n'
                                '}\n'
                                'Rules:\n'
                                '- color_code: suffix after hyphen ONLY. K128-BLK→"BLK", K231-NVY→"NVY". '
                                'Standalone codes like NVY, CRH, HGY use as-is.\n'
                                '- fabric_desc: fiber content from HTS line after | character. '
                                'e.g. "100% COTTON KNIT", "60% COTTON 40% POLY KNIT"\n'
                                '- ex_factory_date / delivery_date: MM/DD/YYYY format\n'
                                '- fob_prices: FOB price per size from the PO price table. '
                                'Regular sizes (XS-2XL) usually same price, '
                                '3XL+ usually higher, Tall (LTLL+) usually different.\n'
                                '- order_unit: total ordered quantity (sum of all sizes)\n'
                                '- Return ONLY JSON, no markdown, no explanation'
                            )
                        }
                    ]
                }]
            )
            text = resp.content[0].text.strip()
            text = re.sub(r'^```json\s*','',text)
            text = re.sub(r'```\s*$','',text).strip()
            po = json.loads(text)
            po['filename'] = f.filename
            results.append(po)
        except Exception as e:
            results.append({'error':str(e),'filename':f.filename})
        time.sleep(3)

    return jsonify({'pos':results})


@app.route('/build-excel', methods=['POST'])
def build_excel():
    try:
        data             = request.json
        file_number      = data.get('file_number','UNKNOWN')
        pos              = data.get('pos',[])
        color_names      = data.get('color_names',{})
        fabric_combos    = data.get('fabric_combos',[])
        sketches         = data.get('sketches',{})
        sewing           = data.get('sewing','')
        printing         = data.get('printing','')
        washing          = data.get('washing','')
        presentation     = data.get('presentation','')
        revision_history = data.get('revision_history',[])

        # EX-FACTORY 날짜 순 정렬
        pos.sort(key=lambda p: parse_date(p.get('ex_factory_date','')) or datetime.date.max)

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # ── File#, Season ──────────────────────────────
        w(ws, 1, 3, file_number)   # C1
        # season은 프론트에서 받는 경우 대비 (현재는 file_number 연도 추출)
        season = data.get('season', '')
        if season:
            w(ws, 2, 3, season)

        # ── 헤더 (col 3) ──────────────────────────────
        def clean(val):
            """직접입력 문자열이 그대로 오면 빈값으로"""
            if val and str(val).strip() == '직접입력':
                return ''
            return val

        w(ws, ROW_SEWING,       3, clean(sewing))
        w(ws, ROW_PRINTING,     3, clean(printing))
        w(ws, ROW_WASHING,      3, clean(washing))
        w(ws, ROW_PRESENTATION, 3, clean(presentation))

        # ── Revision History (row 9+, AC/AE열) ────────
        for i, rev in enumerate(revision_history):
            row = ROW_REVISION_START + i
            w(ws, row, COL_REV_DATE, rev.get('date',''))
            # autoSummary(자동감지) + notes(수동입력) 합쳐서 기록
            auto  = rev.get('autoSummary','')
            notes = rev.get('notes','') or rev.get('description','')
            if auto and notes:
                combined = f"[자동] {auto} / [메모] {notes}"
            elif auto:
                combined = f"[자동] {auto}"
            else:
                combined = notes
            w(ws, row, COL_REV_NOTES, combined)

        # ── Fabric Combination (row 11~14) ─────────────
        for i, fc in enumerate(fabric_combos[:4]):
            row = FABRIC_ROW_START + i
            w(ws, row, FABRIC_COL_COMBO, fc.get('combo', f'C{i+1}'))
            w(ws, row, FABRIC_COL_BCODE, clean(fc.get('bodyCode','')))
            w(ws, row, FABRIC_COL_BDESC, clean(fc.get('bodyDesc','')))
            w(ws, row, FABRIC_COL_TCODE, clean(fc.get('trimCode','')))
            w(ws, row, FABRIC_COL_TDESC, clean(fc.get('trimDesc','')))

        # ── PO 슬롯 (16개) ────────────────────────────
        for slot, po in enumerate(pos[:16]):
            pcs_col = PO_SLOT_COLS[slot]   # 홀수 열 (3,5,7...)
            fob_col = pcs_col + 1          # 짝수 열 (4,6,8...)

            # PO 기본 정보
            w(ws, ROW_PO,         pcs_col, po.get('po_number',''))
            w(ws, ROW_ORDER_UNIT, pcs_col, po.get('order_unit',0) or 0)

            ex  = parse_date(po.get('ex_factory_date',''))
            w(ws, ROW_EX_FACTORY, pcs_col, ex  or po.get('ex_factory_date',''))
            dlv = parse_date(po.get('delivery_date',''))
            w(ws, ROW_DELIVERY,   pcs_col, dlv or po.get('delivery_date',''))

            w(ws, ROW_STYLE,      pcs_col, po.get('style',''))
            w(ws, ROW_COLOR_CODE, pcs_col, po.get('color_code',''))
            w(ws, ROW_COLOR_NAME, pcs_col, color_names.get(po.get('color_code',''),''))

            # Fabric Combo 라벨
            fiber = normalize_fiber(po.get('fabric_desc',''))
            combo_label = ''
            for fc in fabric_combos:
                if normalize_fiber(fc.get('bodyDesc','')) == fiber:
                    combo_label = fc.get('combo','')
                    break
            w(ws, ROW_FABRIC_LBL, pcs_col, combo_label)

            w(ws, ROW_SHIP_TO,   pcs_col, po.get('ship_to',''))
            w(ws, ROW_SHIP_MODE, pcs_col, po.get('ship_mode',''))

            # 사이즈별 PCS + FOB
            sizes     = po.get('sizes',{})
            fob_prices = po.get('fob_prices',{})

            for size, srow in SIZE_ROWS.items():
                qty = sizes.get(size, 0)
                fob = fob_prices.get(size, 0)
                if qty:
                    w(ws, srow, pcs_col, int(qty))
                if fob:
                    w(ws, srow, fob_col, float(fob))

        # ── Row 23(COLOR), 25(SHIP TO) 높이 조정 ────────
        ws.row_dimensions[23].height = 36
        ws.row_dimensions[25].height = 45

        # ── 스케치 이미지 (row4~14 병합셀 중앙 정렬) ──
        # 실제 셀 크기: 3열(13 charwidth) × 11행(16.8pt)
        # column width → px: int(width*7)+5 = 96px, ×3 = 288px
        # row height → px: 16.8pt × (96/72) ≈ 22.4px, ×11 = 246px
        CELL_W_PX = 3 * (int(13 * 7) + 5)   # 288px
        CELL_H_PX = int(11 * 16.8 * 96 / 72) # 246px

        styles_seen = []
        for po in pos:
            sty = po.get('style','')
            if sty and sty not in styles_seen:
                styles_seen.append(sty)

        # Style# 셀에 실제 스타일번호 기재 (row3, SKETCH_COLS 위치)
        for idx, style in enumerate(styles_seen[:4]):
            col = SKETCH_COLS.get(idx, 16)
            w(ws, SKETCH_ROW, col, style)

        # 스케치 이미지 삽입
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        PX2EMU = 9525

        for idx, style in enumerate(styles_seen[:4]):
            if not sketches.get(style):
                continue
            raw = sketches[style]
            if ',' in raw:
                raw = raw.split(',')[1]
            try:
                img_bytes = base64.b64decode(raw)
                img = XLImage(io.BytesIO(img_bytes))

                # 이미지 크기 계산
                target_w = int(5 / 2.54 * 96)   # 5cm ≈ 189px
                target_h = 220
                try:
                    from PIL import Image as PILImage
                    pil = PILImage.open(io.BytesIO(img_bytes))
                    ow, oh = pil.size
                    target_h = int(oh * (target_w / ow))
                except Exception:
                    pass

                # 셀 초과 시 축소
                max_w = CELL_W_PX - 12
                max_h = CELL_H_PX - 12
                if target_w > max_w:
                    target_h = int(target_h * max_w / target_w)
                    target_w = max_w
                if target_h > max_h:
                    target_w = int(target_w * max_h / target_h)
                    target_h = max_h

                img.width  = target_w
                img.height = target_h

                # 중앙 정렬: 셀 왼쪽 상단 기준 오프셋
                x_off = max(0, (CELL_W_PX - target_w) // 2) * PX2EMU
                y_off = max(0, (CELL_H_PX - target_h) // 2) * PX2EMU

                col_0based = SKETCH_COLS.get(idx, 16) - 1  # 0-based
                marker = AnchorMarker(col=col_0based, colOff=x_off, row=3, rowOff=y_off)  # row=3 → Excel row 4
                size   = XDRPositiveSize2D(target_w * PX2EMU, target_h * PX2EMU)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
            except Exception:
                pass

        # ── 저장 & 반환 ───────────────────────────────
        date_str = datetime.date.today().strftime('%Y%m%d')
        filename = f"{file_number}_Order Recap_Team#4_{date_str}.xlsx"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error':str(e),'traceback':traceback.format_exc()}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT',5000))
    app.run(host='0.0.0.0', port=port, debug=False)
